"""Repricing à partir d'un fichier Excel/CSV : compare TON prix au prix
concurrent et recommande « baisser » (tu es trop cher) ou « augmenter »
(tu es bien moins cher, marge disponible). Met en avant les grosses marges.

Aucune clé, aucun scraping : tout se calcule en local depuis ton fichier.
"""
import csv
import io
import re
import unicodedata


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return s.lower().strip()


def _to_float(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    txt = str(v).strip().replace("€", "").replace(" ", "").replace(" ", "")
    if "," in txt and "." in txt:
        txt = txt.replace(".", "").replace(",", ".")
    elif "," in txt:
        txt = txt.replace(",", ".")
    try:
        return float(re.sub(r"[^\d.]", "", txt))
    except ValueError:
        return None


# Détection souple des colonnes (accents/casse ignorés)
COLS = {
    "name": ("produit", "nom", "libelle", "designation", "titre", "name", "article"),
    "ean": ("ean", "gtin", "sku", "reference", "ref", "code"),
    "my_price": ("mon prix", "prix fnac", "prix de vente", "prix vente",
                 "my price", "prix actuel", "prix"),
    "competitor_price": ("prix concurrent", "concurrent", "competitor",
                         "prix marche", "prix mini concurrent", "prix concurrence"),
    "competitor_name": ("site concurrent", "concurrent nom", "source", "vendeur"),
}


def _map_columns(headers: list[str]) -> dict:
    mapping = {}
    norm_headers = [(_norm(h), h) for h in headers]
    for field, keys in COLS.items():
        # priorité aux libellés les plus spécifiques (les plus longs d'abord)
        for key in sorted(keys, key=len, reverse=True):
            for nh, original in norm_headers:
                if key in nh and original not in mapping.values():
                    mapping[field] = original
                    break
            if field in mapping:
                break
    return mapping


def parse_file(content: bytes, filename: str) -> tuple[list[dict], dict]:
    """Lit un CSV ou XLSX -> (lignes, mapping des colonnes)."""
    name = filename.lower()
    rows: list[dict] = []
    if name.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else "" for h in next(it, [])]
        for r in it:
            rows.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
    else:  # CSV
        text = content.decode("utf-8-sig", errors="replace")
        sample = text[:2000]
        delim = ";" if sample.count(";") > sample.count(",") else ","
        reader = csv.DictReader(io.StringIO(text), delimiter=delim)
        headers = reader.fieldnames or []
        rows = list(reader)

    mapping = _map_columns(headers)
    return rows, mapping


def recommend(my_price: float | None, competitor_price: float | None,
              cfg: dict) -> dict:
    """Recommandation d'action pour une ligne."""
    if my_price is None or competitor_price is None or competitor_price <= 0:
        return {"action": "donnees_manquantes", "target": None,
                "gap": None, "gap_percent": None,
                "reason": "Prix manquant ou invalide."}

    gap = round(my_price - competitor_price, 2)
    gap_percent = round(gap / competitor_price * 100, 1)
    undercut = 1 - cfg["undercutPercent"] / 100.0

    if gap_percent > cfg["alignThresholdPercent"]:
        return {"action": "baisser", "target": round(competitor_price * undercut, 2),
                "gap": gap, "gap_percent": gap_percent,
                "reason": f"Tu es {gap_percent:.0f}% plus cher que le concurrent."}
    if gap_percent < -cfg["raiseThresholdPercent"]:
        return {"action": "augmenter", "target": round(competitor_price * undercut, 2),
                "gap": gap, "gap_percent": gap_percent,
                "reason": f"Tu es {abs(gap_percent):.0f}% moins cher : marge disponible."}
    return {"action": "ok", "target": my_price, "gap": gap,
            "gap_percent": gap_percent, "reason": "Prix aligné, rien à changer."}


def analyze(content: bytes, filename: str, cfg: dict) -> dict:
    rows, mapping = parse_file(content, filename)
    missing = [k for k in ("my_price", "competitor_price") if k not in mapping]
    if missing:
        return {"ok": False, "mapping": mapping,
                "error": "Colonnes introuvables : il faut au minimum une colonne "
                         "de TON prix et une colonne PRIX CONCURRENT. "
                         "Télécharge le modèle Excel pour le bon format."}

    results = []
    counts = {"baisser": 0, "augmenter": 0, "ok": 0, "donnees_manquantes": 0}
    for r in rows:
        my_price = _to_float(r.get(mapping.get("my_price")))
        comp = _to_float(r.get(mapping.get("competitor_price")))
        reco = recommend(my_price, comp, cfg)
        counts[reco["action"]] = counts.get(reco["action"], 0) + 1
        results.append({
            "name": str(r.get(mapping.get("name", ""), "") or "").strip(),
            "ean": str(r.get(mapping.get("ean", ""), "") or "").strip(),
            "competitor": str(r.get(mapping.get("competitor_name", ""), "") or "").strip(),
            "my_price": my_price, "competitor_price": comp,
            "action": reco["action"], "target_price": reco["target"],
            "gap": reco["gap"], "gap_percent": reco["gap_percent"],
            "reason": reco["reason"],
        })

    # Grosses marges d'abord (|écart %| décroissant)
    results.sort(key=lambda x: abs(x["gap_percent"]) if x["gap_percent"] is not None
                 else -1, reverse=True)
    return {"ok": True, "count": len(results), "counts": counts,
            "mapping": mapping, "rows": results}


def template_csv() -> bytes:
    """Modèle CSV que l'utilisateur remplit."""
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow(["Produit", "EAN", "Mon prix", "Prix concurrent", "Site concurrent"])
    w.writerow(["Casque Sony WH-1000XM5", "4548736132566", "349.00", "279.99", "Amazon"])
    w.writerow(["TV LG OLED 55 C4", "8806091985702", "1290.00", "1490.00", "Boulanger"])
    w.writerow(["PS5 Slim", "0711719577294", "549.00", "499.00", "Cdiscount"])
    return buf.getvalue().encode("utf-8-sig")
